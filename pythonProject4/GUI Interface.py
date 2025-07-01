import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import subprocess
from datetime import datetime, timedelta
import os
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- SQL Config ---
DB_SERVER = 'SBNC-sql'
DB_NAME = 'NGProd'
ODBC_DRIVER = 'ODBC Driver 17 for SQL Server'

def get_engine():
    conn_str = (
        f"mssql+pyodbc://@{DB_SERVER}/{DB_NAME}"
        f"?driver={ODBC_DRIVER.replace(' ', '+')}&Trusted_Connection=yes"
    )
    return create_engine(conn_str)

def export_to_excel(df, output_path, sheet_name="Results"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    n_rows, n_cols = df.shape
    last_col = get_column_letter(n_cols)
    table_ref = f"A1:{last_col}{n_rows + 1}"
    tab = Table(displayName=sheet_name[:31].replace(" ", ""), ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(10, min(max_len + 2, 40))
    wb.save(output_path)

def run_main_template_query(lower_date_str, upper_date_str):
    try:
        lower_date = datetime.strptime(lower_date_str, "%Y-%m-%d")
        upper_date = datetime.strptime(upper_date_str, "%Y-%m-%d")
    except ValueError:
        messagebox.showerror("Date Format Error", "Dates must be in YYYY-MM-DD format.")
        return

    upper_date_adj = upper_date - timedelta(days=1)
    lower_str_sql = lower_date.strftime("%Y%m%d")
    upper_str_sql = upper_date_adj.strftime("%Y%m%d")
    file_date_range = f"{lower_date.strftime('%Y-%m-%d')} to {upper_date.strftime('%Y-%m-%d')}"
    output_file = fr'C:\Reports\Provider Prod Data Pulls\Prov Prod Data {file_date_range}.xlsx'

    sql_query = f"""
    SELECT 
        e.category,
        r.week_start_date, 
        r.week_end_date,
        c.create_timestamp AS [Date Appt Was Created],
        e.prevent_appts_ind AS [Prevent Appointments?], 
        c.duration,
        t.template,
        y.description AS [Provider]
    FROM 
        template_members c
    INNER JOIN 
        appt_templates t ON t.appt_template_id = c.appt_template_id
    INNER JOIN 
        categories e ON e.category_id = c.category_id
    INNER JOIN 
        resource_templates r ON r.appt_template_id = t.appt_template_id
    INNER JOIN 
        resources y ON y.resource_id = r.resource_id
    WHERE 
        CONVERT(VARCHAR, r.week_start_date, 112) BETWEEN '{lower_str_sql}' AND '{upper_str_sql}'
    ORDER BY 
        r.week_start_date DESC;
    """

    try:
        engine = get_engine()
        df = pd.read_sql(sql_query, engine)
        if df.empty:
            append_output("⚠️ No results found.")
        else:
            export_to_excel(df, output_file, "Template Schedule")
            append_output(f"✅ File exported to {output_file}")
    except Exception as e:
        append_output(f"❌ SQL ERROR: {e}")

# --- R Script Runners ---
def run_r_script(script_name, args=None):
    if args is None:
        args = []
    r_exe = r"C:/Users/calvin.baker_SBNC/AppData/Local/Programs/R/R-4.3.2/bin/Rscript.exe"
    r_script_path = rf"\\SBNC-file1\users\calvin.baker_SBNC\Documents\R Scripts\Prov Prod\Auto_R_Filtering\{script_name}"
    try:
        result = subprocess.run([r_exe, r_script_path] + args, capture_output=True, text=True)
        append_output("📤 STDOUT:\n" + (result.stdout or "[no output]"))
        if result.stderr:
            append_output("❌ STDERR:\n" + result.stderr)
        append_output(f"🔚 Exit Code: {result.returncode}")
    except Exception as e:
        append_output(f"❌ Subprocess error: {e}")

# --- GUI Output ---
def append_output(text):
    output_box.config(state="normal")
    output_box.insert(tk.END, text + "\n")
    output_box.see(tk.END)
    output_box.config(state="disabled")

# --- Task Handler ---
def run_task():
    task_type = main_choice_var.get()
    if task_type == "Provider Productivity File Pull":
        lower = lower_entry.get().strip()
        upper = upper_entry.get().strip()
        if not lower or not upper:
            messagebox.showerror("Input Error", "Both dates required.")
            return
        run_main_template_query(lower, upper)
    else:
        script_choice = sub_choice_var.get()
        if script_choice == "Incentive Calculation":
            pay_period = pay_period_entry.get().strip()
            if not pay_period:
                messagebox.showerror("Missing Date", "Enter pay period start (YYYY-MM-DD).")
                return
            run_r_script("Incentive_Calc.R", [pay_period])
        elif script_choice == "4 Week Interval Workbook Only":
            run_r_script("Run_4Week.R")
        elif script_choice == "ISO Week Workbook Only":
            run_r_script("Run_IsoWeek.R")
        else:
            messagebox.showwarning("Choose Option", "Select a script option.")

def on_main_choice_change(*_):
    if main_choice_var.get() == "Provider Productivity File Pull":
        lower_label.grid()
        lower_entry.grid()
        upper_label.grid()
        upper_entry.grid()
        sub_choice_label.grid_remove()
        sub_choice_menu.grid_remove()
        pay_period_label.grid_remove()
        pay_period_entry.grid_remove()
    else:
        lower_label.grid_remove()
        lower_entry.grid_remove()
        upper_label.grid_remove()
        upper_entry.grid_remove()
        sub_choice_label.grid()
        sub_choice_menu.grid()

def on_sub_choice_change(*_):
    if sub_choice_var.get() == "Incentive Calculation":
        pay_period_label.grid()
        pay_period_entry.grid()
    else:
        pay_period_label.grid_remove()
        pay_period_entry.grid_remove()

# --- GUI Layout ---
root = tk.Tk()
root.title("📊 SBNC Productivity Tool")
root.geometry("850x600")

main_choice_var = tk.StringVar()
main_choice_var.trace_add("write", on_main_choice_change)
main_menu = ttk.Combobox(root, textvariable=main_choice_var, state="readonly", width=50)
main_menu['values'] = ["Provider Productivity File Pull", "Run R Filtering Sequence"]
main_menu.current(0)

tk.Label(root, text="Choose Task:").grid(row=0, column=0, padx=10, pady=10, sticky="w")
main_menu.grid(row=0, column=1, padx=10, pady=10)

lower_label = tk.Label(root, text="Lower Limit Date (YYYY-MM-DD):")
lower_entry = tk.Entry(root, width=20)
lower_label.grid(row=1, column=0, sticky="w", padx=10)
lower_entry.grid(row=1, column=1, pady=5)

upper_label = tk.Label(root, text="Upper Limit Date (Last Sunday, YYYY-MM-DD):")
upper_entry = tk.Entry(root, width=20)
upper_label.grid(row=2, column=0, sticky="w", padx=10)
upper_entry.grid(row=2, column=1, pady=5)

sub_choice_var = tk.StringVar()
sub_choice_var.trace_add("write", on_sub_choice_change)
sub_choice_label = tk.Label(root, text="Choose R Script:")
sub_choice_menu = ttk.Combobox(root, textvariable=sub_choice_var, state="readonly", width=50)
sub_choice_menu['values'] = ["Incentive Calculation", "4 Week Interval Workbook Only", "ISO Week Workbook Only"]
sub_choice_label.grid_remove()
sub_choice_menu.grid_remove()

pay_period_label = tk.Label(root, text="Pay Period Start (YYYY-MM-DD):")
pay_period_entry = tk.Entry(root, width=20)
pay_period_label.grid_remove()
pay_period_entry.grid_remove()

tk.Button(root, text="▶️ Run", command=run_task).grid(row=4, column=1, pady=10)

output_box = scrolledtext.ScrolledText(root, wrap=tk.WORD, width=100, height=20, state="disabled")
output_box.grid(row=5, column=0, columnspan=2, padx=10, pady=10)

root.mainloop()

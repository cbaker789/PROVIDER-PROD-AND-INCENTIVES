from sqlalchemy import create_engine
import pandas as pd
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- CONFIGURATION ---
DB_SERVER = 'SBNC-sql'
DB_NAME = 'NGProd'
ODBC_DRIVER = 'ODBC Driver 17 for SQL Server'


# --- Prompt Date Input ---
def prompt_date(prompt_text):
    while True:
        date_str = input(f"{prompt_text} (YYYYMMDD): ")
        try:
            return datetime.strptime(date_str, "%Y%m%d")
        except ValueError:
            print("❌ Invalid format. Please use YYYYMMDD.")


# --- Format Excel Sheet as Table ---
def format_sheet_as_table(sheet, df, table_name):
    if df.empty:
        return
    n_rows, n_cols = df.shape
    last_col = get_column_letter(n_cols)
    table_ref = f"A1:{last_col}{n_rows + 1}"
    tab = Table(displayName=table_name[:31].replace(" ", ""), ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    sheet.add_table(tab)

    for col_cells in sheet.columns:
        col_letter = col_cells[0].column_letter
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col_cells)
        sheet.column_dimensions[col_letter].width = max(10, min(max_len + 2, 40))


# --- Export to Excel ---
def export_to_excel(df, output_path, sheet_name="Results"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    format_sheet_as_table(ws, df, sheet_name)
    wb.save(output_path)
    print(f"[✓] Excel exported to: {output_path}")


# --- Connection Builder ---
def get_engine():
    conn_str = (
        f"mssql+pyodbc://@{DB_SERVER}/{DB_NAME}"
        f"?driver={ODBC_DRIVER.replace(' ', '+')}&Trusted_Connection=yes"
    )
    return create_engine(conn_str)


# --- Query Runner ---
def run_query_and_export(sql_query, file_name, sheet_name="Results"):
    try:
        engine = get_engine()
        print("[...] Running SQL Query...")
        df = pd.read_sql(sql_query, engine)

        if df.empty:
            print("⚠️ No results found for the selected date range.")
        else:
            export_to_excel(df, file_name, sheet_name)
    except Exception as e:
        print(f"[!] ERROR: {e}")


# --- Main Template Query ---
def run_main_template_query():
    lower_date = prompt_date("Enter the LOWER limit date")
    upper_date = prompt_date("Enter the date of the last SUNDAY that passed (Upper limit)")
    upper_date_adj = upper_date - timedelta(days=1)

    lower_str_sql = lower_date.strftime("%Y%m%d")
    upper_str_sql = upper_date_adj.strftime("%Y%m%d")
    file_date_range = f"{lower_date.strftime('%Y-%m-%d')} to {upper_date.strftime('%Y-%m-%d')}"
    output_file = fr'C:\Reports\Provider Prod Data Pulls\Prov Prod Data {file_date_range}.xlsx'

    sql_query = f"""
    SELECT 
        e.category,
        r.week_start_date, 
        r.week_end_date,
        c.create_timestamp AS [Date Appt Was Created],
        e.prevent_appts_ind AS [Prevent Appointments?], 
        c.duration,
        t.template,
        y.description AS [Provider]
    FROM 
        template_members c
    INNER JOIN 
        appt_templates t ON t.appt_template_id = c.appt_template_id
    INNER JOIN 
        categories e ON e.category_id = c.category_id
    INNER JOIN 
        resource_templates r ON r.appt_template_id = t.appt_template_id
    INNER JOIN 
        resources y ON y.resource_id = r.resource_id
    WHERE 
        CONVERT(VARCHAR, r.week_start_date, 112) BETWEEN '{lower_str_sql}' AND '{upper_str_sql}'
    ORDER BY 
        r.week_start_date DESC;
    """

    run_query_and_export(sql_query, output_file, "Template Schedule")

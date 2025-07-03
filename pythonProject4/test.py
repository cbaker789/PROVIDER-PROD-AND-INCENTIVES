import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import subprocess
from datetime import datetime, timedelta
import os
import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# --- SQL Config ---
DB_SERVER = 'SBNC-sql'
DB_NAME = 'NGProd'
ODBC_DRIVER = 'ODBC Driver 17 for SQL Server'

def get_engine():
    conn_str = (
        f"mssql+pyodbc://@{DB_SERVER}/{DB_NAME}"
        f"?driver={ODBC_DRIVER.replace(' ', '+')}&Trusted_Connection=yes"
    )
    return create_engine(conn_str)

def export_to_excel(df, output_path, sheet_name="Results"):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    ws = wb[sheet_name]
    n_rows, n_cols = df.shape
    last_col = get_column_letter(n_cols)
    table_ref = f"A1:{last_col}{n_rows + 1}"
    tab = Table(displayName=sheet_name[:31].replace(" ", ""), ref=table_ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max(10, min(max_len + 2, 40))
    wb.save(output_path)

def run_main_template_query(lower_date_str, upper_date_str):
    try:
        lower_date = datetime.strptime(lower_date_str, "%Y-%m-%d")
        upper_date = datetime.strptime(upper_date_str, "%Y-%m-%d")
    except ValueError:
        messagebox.showerror("Date Format Error", "Dates must be in YYYY-MM-DD format.")
        return

    upper_date_adj = upper_date - timedelta(days=1)
    lower_str_sql = lower_date.strftime("%Y%m%d")
    upper_str_sql = upper_date_adj.strftime("%Y%m%d")
    file_date_range = f"{lower_date.strftime('%Y-%m-%d')} to {upper_date.strftime('%Y-%m-%d')}"
    output_file = fr'C:\Reports\Provider Prod Data Pulls\Prov Prod Data {file_date_range}.xlsx'

    sql_query = f"""
    SELECT 
        e.category,
        r.week_start_date, 
        r.week_end_date,
        c.create_timestamp AS [Date Appt Was Created],
        e.prevent_appts_ind AS [Prevent Appointments?], 
        c.duration,
        t.template,
        y.description AS [Provider]
    FROM 
        template_members c
    INNER JOIN 
        appt_templates t ON t.appt_template_id = c.appt_template_id
    INNER JOIN 
        categories e ON e.category_id = c.category_id
    INNER JOIN 
        resource_templates r ON r.appt_template_id = t.appt_template_id
    INNER JOIN 
        resources y ON y.resource_id = r.resource_id
    WHERE 
        CONVERT(VARCHAR, r.week_start_date, 112) BETWEEN '{lower_str_sql}' AND '{upper_str_sql}'
    ORDER BY 
        r.week_start_date DESC;
    """

    try:
        engine = get_engine()
        df = pd.read_sql(sql_query, engine)
        if df.empty:
            append_output("⚠️ No results found.")
        else:
            export_to_excel(df, output_file, "Template Schedule")
            append_output(f"✅ File exported to {output_file}")
    except Exception as e:
        append_output(f"❌ SQL ERROR: {e}")

def run_r_script(script_name, args=None):
    if args is None:
        args = []
    r_exe = r"C:/Users/calvin.baker_SBNC/AppData/Local/Programs/R/R-4.3.2/bin/Rscript.exe"
    r_script_path = rf"\\SBNC-file1\users\calvin.baker_SBNC\Documents\R Scripts\Prov Prod\Auto_R_Filtering\{script_name}"
    try:
        result = subprocess.run([r_exe, r_script_path] + args, capture_output=True, text=True)
        append_output("📤 STDOUT:\n" + (result.stdout or "[no output]"))
        if result.stderr:
            append_output("❌ STDERR:\n" + result.stderr)
        append_output(f"🔚 Exit Code: {result.returncode}")
    except Exception as e:
        append_output(f"❌ Subprocess error: {e}")

def append_output(text):
    output_box.config(state="normal")
    output_box.insert(tk.END, text + "\n")
    output_box.see(tk.END)
    output_box.config(state="disabled")

def run_task():
    task_type = main_choice_var.get()
    if task_type == "Provider Productivity File Pull":
        lower = lower_entry.get().strip()
        upper = upper_entry.get().strip()
        if not lower or not upper:
            messagebox.showerror("Input Error", "Both dates required.")
            return
        run_main_template_query(lower, upper)
    else:
        script_choice = sub_choice_var.get()
        if script_choice == "Incentive Calculation":
            pay_period = pay_period_entry.get().strip()
            if not pay_period:
                messagebox.showerror("Missing Date", "Enter pay period start (YYYY-MM-DD).")
                return
            run_r_script("Incentive_Calc.R", [pay_period])
        elif script_choice == "4 Week Interval Workbook Only":
            run_r_script("Run_4Week.R")
        elif script_choice == "ISO Week Workbook Only":
            run_r_script("Run_IsoWeek.R")
        elif script_choice == "ISO Week bY PROVIDER":
            run_r_script("ISO_Week_Split_By_Provider.R")
        else:
            messagebox.showwarning("Choose Option", "Select a script option.")

def on_main_choice_change(*_):
    if main_choice_var.get() == "Provider Productivity File Pull":
        frame_dates.grid()
        frame_rfilter.grid_remove()
    else:
        frame_dates.grid_remove()
        frame_rfilter.grid()

def on_sub_choice_change(*_):
    if sub_choice_var.get() == "Incentive Calculation":
        pay_period_label.grid()
        pay_period_entry.grid()
    else:
        pay_period_label.grid_remove()
        pay_period_entry.grid_remove()

# --- GUI ---
root = tk.Tk()
root.title("📊 SBNC Productivity Tool")
root.geometry("950x650")
root.configure(bg="#f4f4f4")

style = ttk.Style(root)
style.theme_use("clam")

main_choice_var = tk.StringVar()
main_choice_var.trace_add("write", on_main_choice_change)

sub_choice_var = tk.StringVar()
sub_choice_var.trace_add("write", on_sub_choice_change)

# Task Selection Frame
frame_top = ttk.LabelFrame(root, text="Select Task", padding=(15, 10))
frame_top.grid(row=0, column=0, columnspan=2, padx=15, pady=10, sticky="ew")

ttk.Label(frame_top, text="Choose Task:").grid(row=0, column=0, padx=5, pady=5, sticky="w")
main_menu = ttk.Combobox(frame_top, textvariable=main_choice_var, state="readonly", width=50)
main_menu['values'] = ["Provider Productivity File Pull", "Run R Filtering Sequence"]
main_menu.current(0)
main_menu.grid(row=0, column=1, padx=5, pady=5, sticky="w")

# Date Entry Frame
frame_dates = ttk.LabelFrame(root, text="Date Range", padding=(15, 10))
frame_dates.grid(row=1, column=0, columnspan=2, padx=15, pady=5, sticky="ew")

lower_label = ttk.Label(frame_dates, text="Lower Limit Date (YYYY-MM-DD):")
lower_entry = ttk.Entry(frame_dates, width=20)
lower_label.grid(row=0, column=0, sticky="w", padx=5, pady=5)
lower_entry.grid(row=0, column=1, pady=5, sticky="w")

upper_label = ttk.Label(frame_dates, text="Upper Limit Date (Last Sunday, YYYY-MM-DD):")
upper_entry = ttk.Entry(frame_dates, width=20)
upper_label.grid(row=1, column=0, sticky="w", padx=5, pady=5)
upper_entry.grid(row=1, column=1, pady=5, sticky="w")

# R Script Frame
frame_rfilter = ttk.LabelFrame(root, text="R Filtering Options", padding=(15, 10))
frame_rfilter.grid(row=2, column=0, columnspan=2, padx=15, pady=5, sticky="ew")

sub_choice_label = ttk.Label(frame_rfilter, text="Choose R Script:")
sub_choice_menu = ttk.Combobox(frame_rfilter, textvariable=sub_choice_var, state="readonly", width=50)
sub_choice_label.grid(row=0, column=0, sticky="w", padx=5, pady=5)
sub_choice_menu.grid(row=0, column=1, padx=5, pady=5)
sub_choice_menu['values'] = ["Incentive Calculation", "4 Week Interval Workbook Only", "ISO Week Workbook Only", "ISO Week bY PROVIDER"]

pay_period_label = ttk.Label(frame_rfilter, text="Pay Period Start (YYYY-MM-DD):")
pay_period_entry = ttk.Entry(frame_rfilter, width=20)
pay_period_label.grid(row=1, column=0, sticky="w", padx=5, pady=5)
pay_period_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")

# Run Button
ttk.Button(root, text="▶️ Run Task", command=run_task).grid(row=3, column=1, pady=10, sticky="e", padx=15)

# Output Frame
output_frame = ttk.LabelFrame(root, text="Console Output", padding=(10, 10))
output_frame.grid(row=4, column=0, columnspan=2, padx=15, pady=10, sticky="nsew")
root.grid_rowconfigure(4, weight=1)
root.grid_columnconfigure(1, weight=1)

output_box = scrolledtext.ScrolledText(output_frame, wrap=tk.WORD, width=120, height=20, font=("Courier New", 10), state="disabled", bg="#fcfcfc", relief="solid", borderwidth=1)
output_box.pack(fill="both", expand=True)

# Initialize layout visibility
on_main_choice_change()
on_sub_choice_change()

root.mainloop()
